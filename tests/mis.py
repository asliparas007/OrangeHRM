import undetected_chromedriver as uc
from selenium import webdriver
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import time

from openpyxl import load_workbook

workbook = load_workbook(r'C:\Users\paras\Desktop\Test1.xlsx')
sheet = workbook.active

username = sheet.cell(row=1,column=1).value
print(username)

# for row in sheet.iter_rows(values_only=True):
#     print(row)


