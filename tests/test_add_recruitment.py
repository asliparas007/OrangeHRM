import pytest
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from PageObjects.AddRecruitmentPage import AddRecruitment
from openpyxl import load_workbook
#
workbook = load_workbook(r'C:\Users\paras\Desktop\Test1.xlsx')
sheet = workbook.active

firstname = sheet.cell(row=1,column=1).value
lastname = sheet.cell(row=1,column=2).value
email = sheet.cell(row=1,column=3).value


def test_recruitment(valid_user_session):
    driver = valid_user_session
    # driver.get_screenshot_as_file("screenshot.png")
    recruitment_option = AddRecruitment(driver)
    recruitment_option.click_recruitment_option()
    recruitment_option.click_add_button()
    recruitment_option.add_candidate(firstname,lastname,email)
    #-------load and select from custom drop down
    recruitment_option.add_vacancy()
    #------------------------
    

